from django.db import transaction
from django.shortcuts import render, get_object_or_404
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from django.utils import timezone
from rest_framework import status
from rest_framework.decorators import api_view
from rest_framework.response import Response
from .models import Expense, ExpenseCategory

# تجميع الاستيرادات
from .models import Order, GiftCard, Payment, OrderItem
from .serializers import OrderSerializer, GiftCardSerializer, OrderItemSerializer
from django.db.models import Sum, Q
import datetime
from .models import Order, GiftCard, Payment, OrderItem, ShiftReport #
from django.db.models.functions import Coalesce
from django.db.models import FloatField
from .models import Order, GiftCard, Payment, OrderItem, ShiftReport, Expense
import csv
from django.http import JsonResponse, HttpResponse
from django.db.models.functions import TruncDate
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from django.utils.dateparse import parse_date
from django.db.models import Q
from .models import Order
# --- 1. دالة عرض لوحة تحكم الإدارة ---
def admin_dashboard(request):
    return render(request, 'admin_dashboard.html')

# --- 2. دالة عرض صفحة الإدارة ---
def orders_management(request):
    return render(request, 'orders.html')

# --- 3. كلاس الـ API الخاص بالطلبات ---
class OrderViewSet(viewsets.ModelViewSet):
    queryset = Order.objects.all().order_by('-created_at')
    serializer_class = OrderSerializer

    def get_queryset(self):
        queryset = super().get_queryset()
        customer_id = self.request.query_params.get('customer')
        if customer_id:
            queryset = queryset.filter(customer_id=customer_id)
        return queryset

    # تسجيل الوقت تلقائياً عند تغيير الحالة في الورشة
    def perform_update(self, serializer):
        instance = serializer.save()
        if instance.status == 'PROCESSING' and not instance.processing_start_at:
            instance.processing_start_at = timezone.now()
            instance.save()
        elif instance.status == 'READY' and not instance.ready_at:
            instance.ready_at = timezone.now()
            instance.save()

    @action(detail=False, methods=['get'], url_path='pending')
    def pending_orders(self, request):
        try:
            customer_id = request.query_params.get('customer')
            if not customer_id:
                return Response({"error": "الرجاء تحديد العميل"}, status=status.HTTP_400_BAD_REQUEST)

            orders = Order.objects.filter(customer_id=customer_id, is_void=False).prefetch_related('payments')
            result = []
            for order in orders:
                total_paid = sum(float(p.amount or 0) for p in order.payments.all())
                total_order = float(order.total_amount or 0)
                remaining = total_order - total_paid
                if remaining > 0.01:
                    result.append({
                        "id": order.id,
                        "total_amount": round(remaining, 2),
                        "original_total": total_order,
                        "date": order.created_at.strftime('%d/%m/%Y')
                    })
            return Response(result, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['post'], url_path='pay-receipt')
    def pay_receipt(self, request):
        try:
            data = request.data
            payment_method = data.get('payment_method')
            orders_to_pay = data.get('orders', [])
            total_amount_to_pay = sum(float(item['amount']) for item in orders_to_pay)

            with transaction.atomic():
                if payment_method == 'WALLET':
                    first_order_id = orders_to_pay[0]['order_id'] if orders_to_pay else None
                    if first_order_id:
                        order_for_customer = Order.objects.get(id=first_order_id)
                        customer = order_for_customer.customer
                        if not customer:
                            return Response({"error": "لا يمكن الدفع بالمحفظة لعميل نقدي."}, status=status.HTTP_400_BAD_REQUEST)
                        current_balance = float(customer.wallet_balance or 0)
                        if current_balance < total_amount_to_pay:
                            return Response({"error": "رصيد المحفظة غير كافٍ"}, status=status.HTTP_400_BAD_REQUEST)
                        customer.wallet_balance = current_balance - total_amount_to_pay
                        customer.save()

                elif payment_method == 'GIFT_CARD':
                    gift_card_number = data.get('gift_card_number')
                    gift_card = GiftCard.objects.get(card_number=gift_card_number, is_active=True)
                    if float(gift_card.balance) < total_amount_to_pay:
                        return Response({"error": "رصيد بطاقة الهدية غير كافٍ"}, status=status.HTTP_400_BAD_REQUEST)
                    gift_card.balance = float(gift_card.balance) - total_amount_to_pay
                    if gift_card.balance <= 0:
                        gift_card.is_active = False
                    gift_card.save()

                for item in orders_to_pay:
                    order = Order.objects.get(id=item['order_id'])
                    Payment.objects.create(order=order, amount=item['amount'], method=payment_method)
            return Response({"message": "تم السداد بنجاح"}, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

# --- 4. كلاس الـ API الخاص بالقطع والبطاقات ---
class OrderItemViewSet(viewsets.ModelViewSet):
    queryset = OrderItem.objects.all()
    serializer_class = OrderItemSerializer

class GiftCardViewSet(viewsets.ModelViewSet):
    queryset = GiftCard.objects.all().order_by('-created_at')
    serializer_class = GiftCardSerializer

    # 🌟 التعديل السحري: إجبار السيرفر يرسل اسم العميل ورقم الملف مع لستة البطاقات 🌟
    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        serializer = self.get_serializer(queryset, many=True)
        
        # إضافة الاسم ورقم الملف لكل بطاقة قبل إرسالها للواجهة الأمامية
        modified_data = []
        for index, data in enumerate(serializer.data):
            card_obj = queryset[index]
            data['customer_name'] = card_obj.customer.name if card_obj.customer else "عميل نقدي"
            # حاول سحب رقم الملف (file_number) وإذا غير موجود اسحب الـ id
            if card_obj.customer:
                data['customer_file_number'] = getattr(card_obj.customer, 'file_number', card_obj.customer.id)
            else:
                data['customer_file_number'] = "---"
            modified_data.append(data)
            
        return Response(modified_data)

    @action(detail=False, methods=['get'], url_path='check-balance')
    def check_balance(self, request):
        card_number = request.query_params.get('card_number')
        try:
            gift_card = GiftCard.objects.get(card_number=card_number, is_active=True)
            return Response({
                "id": gift_card.id, "card_number": gift_card.card_number,
                "balance": float(gift_card.balance),
                "customer_name": gift_card.customer.name if gift_card.customer else "عميل نقدي",
                "customer_file_number": getattr(gift_card.customer, 'file_number', gift_card.customer.id) if gift_card.customer else "---"
            }, status=status.HTTP_200_OK)
        except GiftCard.DoesNotExist:
            return Response({"error": "البطاقة غير موجودة"}, status=status.HTTP_404_NOT_FOUND)
# --- 5. دالة طباعة الفاتورة ---
def print_order_receipt(request, order_id):
    order = get_object_or_404(Order, id=order_id)
    total_qty = sum(item.quantity for item in order.items.all())
    total_paid = sum(p.amount for p in order.payments.all())
    is_paid = total_paid >= order.total_amount
    prev_balance = order.customer.wallet_balance if order.customer else None

    context = {
        'order': order,
        'total_qty': total_qty,
        'is_paid': is_paid,
        'prev_balance': prev_balance,
    }
    return render(request, 'receipt.html', context)

# --- 6. دالة عرض شاشة الورشة ---
def workshop_view(request):
    return render(request, 'workshop.html')
# API لجلب التصنيفات (لتعريف القائمة المنسدلة في النافذة)
@api_view(['GET'])
def expense_categories_list(request):
    categories = ExpenseCategory.objects.all().values('id', 'name')
    return Response(list(categories))

# API لحفظ المصروف الجديد
@api_view(['POST'])
def save_expense(request):
    try:
        data = request.data
        Expense.objects.create(
            date=data.get('date'),
            category_id=data.get('category'),
            description=data.get('description'),
            total_amount=data.get('total_amount'),
            paid_amount=data.get('paid_amount'),
            payment_method=data.get('payment_method'),
            notes=data.get('notes')
        )
        return Response({"message": "Saved successfully"}, status=status.HTTP_201_CREATED)
    except Exception as e:
        return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

def items_management(request):
    # هذه الدالة تخبر دجانغو أن يعرض ملف items_management.html
    return render(request, 'items_management.html')

def staff_management(request):
    return render(request, 'staff_management.html')

from django.db.models.functions import Coalesce # 🌟 أضف هذا الاستيراد في الأعلى

from django.db import transaction, models # 🌟 ضروري جداً
from django.shortcuts import render, get_object_or_404
from rest_framework import viewsets, status
from rest_framework.decorators import action, api_view
from rest_framework.response import Response
from django.utils import timezone
from django.db.models import Sum, Q, FloatField # 🌟 أضفنا FloatField
from django.db.models.functions import Coalesce # 🌟 لمنع الـ None
import datetime

# استيراد الموديلات
from .models import Order, GiftCard, Payment, OrderItem, ShiftReport, Expense
from django.db.models import Sum, Count, Q
from datetime import datetime, timedelta

@api_view(['GET', 'POST'])
def close_shift_api(request):
    if request.method == 'GET':
        try:
            today = timezone.now().replace(hour=0, minute=0, second=0, microsecond=0)
            orders_today = Order.objects.filter(created_at__gte=today, is_void=False)
            
            # 🌟 استخدام Coalesce يضمن أن القيمة ستكون 0.0 بدلاً من None 🌟
            total_sales = float(orders_today.aggregate(s=Coalesce(Sum('total_amount'), 0.0, output_field=FloatField()))['s'])
            
            payments = Payment.objects.filter(order__in=orders_today)
            cash_p = float(payments.filter(method='CASH').aggregate(s=Coalesce(Sum('amount'), 0.0, output_field=FloatField()))['s'])
            card_p = float(payments.filter(method='CARD').aggregate(s=Coalesce(Sum('amount'), 0.0, output_field=FloatField()))['s'])
            wallet_p = float(payments.filter(method='WALLET').aggregate(s=Coalesce(Sum('amount'), 0.0, output_field=FloatField()))['s'])
            
            expenses = float(Expense.objects.filter(created_at__gte=today, payment_method='CASH').aggregate(s=Coalesce(Sum('paid_amount'), 0.0, output_field=FloatField()))['s'])
            deleted = float(Order.objects.filter(created_at__gte=today, is_void=True).aggregate(s=Coalesce(Sum('total_amount'), 0.0, output_field=FloatField()))['s'])

            shift_data = {
                "sessionId": "SES" + timezone.now().strftime("%Y%m%d%H%M%S"),
                "totalSales": total_sales,
                "cashPayment": cash_p,
                "cardPayment": card_p,
                "customerBalance": wallet_p,
                "creditSales": max(0, total_sales - (cash_p + card_p + wallet_p)),
                "expensesCash": expenses,
                "deletedOrders": deleted,
                "totalCollection": (cash_p + card_p) - expenses,
                "cashier": request.user.username if request.user.is_authenticated else "Admin"
            }
            return Response(shift_data, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({"error": f"Backend Error: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    elif request.method == 'POST':
        try:
            data = request.data
            sys_data = data.get('system_data', {})
            ShiftReport.objects.create(
                session_id=sys_data.get('sessionId'),
                cashier=request.user if request.user.is_authenticated else None,
                total_sales=sys_data.get('totalSales', 0),
                actual_cash=data.get('actual_cash', 0),
                actual_card=data.get('actual_card', 0),
                expected_cash=data.get('expected_cash', 0),
                difference=data.get('difference', 0),
                notes=data.get('notes', '')
            )
            return Response({"message": "Saved Success"}, status=status.HTTP_201_CREATED)
        except Exception as e:
            return Response({"error": f"Save Error: {str(e)}"}, status=status.HTTP_400_BAD_REQUEST)

def shift_reports_list(request):
    reports = ShiftReport.objects.all().order_by('-closed_at')
    return render(request, 'shift_reports_list.html', {'reports': reports})

# ==========================================
# 📊 قسم التقارير (المربوط بقاعدة البيانات الحقيقية) 📊
# ==========================================
from django.db.models.functions import TruncDate

def orders_report_page(request):
    return render(request, 'orders_report.html', {'categories': []})

def api_orders_report_data(request):
    start_date_str = request.GET.get('start', '2000-01-01')
    end_date_str = request.GET.get('end', '2100-01-01')
    search_query = request.GET.get('search', '').strip().lower()

    try:
        start_date = datetime.strptime(start_date_str, "%Y-%m-%d").date()
        end_date = datetime.strptime(end_date_str, "%Y-%m-%d").date()
    except ValueError:
        return JsonResponse({'error': 'Invalid date format'}, status=400)

    base_query = Order.objects.filter(
        created_at__date__gte=start_date,
        created_at__date__lte=end_date,
        is_void=False
    )

    if search_query:
        base_query = base_query.filter(id__icontains=search_query) | base_query.filter(customer__name__icontains=search_query)

    kpi_data = base_query.aggregate(total_rev=Sum('total_amount'), total_ord=Count('id'))
    total_revenue = float(kpi_data['total_rev'] or 0.00)
    total_orders = kpi_data['total_ord'] or 0

    express_orders_count = base_query.filter(items__is_express=True).distinct().count()
    express_pct = (express_orders_count / total_orders * 100) if total_orders > 0 else 0
    total_items = OrderItem.objects.filter(order__in=base_query).aggregate(Sum('quantity'))['quantity__sum'] or 0

    days_diff = (end_date - start_date).days + 1
    daily_average = total_revenue / days_diff if days_diff > 0 else total_revenue

    daily_stats = base_query.annotate(date=TruncDate('created_at')).values('date').annotate(
        daily_rev=Sum('total_amount'), daily_ord=Count('id', distinct=True)
    ).order_by('date')

    express_stats = OrderItem.objects.filter(order__in=base_query, is_express=True).annotate(
        date=TruncDate('order__created_at')
    ).values('date').annotate(daily_exp=Sum('quantity')).order_by('date')
    express_dict = {str(item['date']): item['daily_exp'] for item in express_stats}

    chart_labels, chart_revenue, chart_orders, chart_express = [], [], [], []
    for stat in daily_stats:
        date_str = str(stat['date'])
        chart_labels.append(stat['date'].strftime('%d/%m'))
        chart_revenue.append(float(stat['daily_rev'] or 0))
        chart_orders.append(stat['daily_ord'])
        chart_express.append(express_dict.get(date_str, 0))

    table_data = []
    for order in base_query.order_by('-created_at'):
        items_qty = order.items.aggregate(Sum('quantity'))['quantity__sum'] or 0
        customer_name = order.customer.name if order.customer else 'Cash Customer'
        table_data.append({
            'date': order.created_at.strftime('%d %b %Y'),
            'order_no': f"ORD-{order.id:04d}",
            'customer': customer_name,
            'items': items_qty,
            'amount': float(order.total_amount),
            'status': order.get_status_display()
        })

    return JsonResponse({
        'kpis': {
            'total_revenue': round(total_revenue, 2), 'total_orders': total_orders,
            'express_orders': express_orders_count, 'express_percentage': round(express_pct, 1),
            'total_items': total_items, 'daily_average': round(daily_average, 2)
        },
        'chart': {
            'labels': chart_labels, 'revenue': chart_revenue, 'orders': chart_orders, 'items': chart_express
        },
        'table': table_data
    })

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

def export_orders_csv(request):
    """ تصدير ملف إكسل حقيقي واحترافي (.xlsx) """
    start_date_str = request.GET.get('start', '2000-01-01')
    end_date_str = request.GET.get('end', '2100-01-01')
    search_query = request.GET.get('search', '').strip().lower()
    
    try:
        start_date = datetime.strptime(start_date_str, "%Y-%m-%d").date()
        end_date = datetime.strptime(end_date_str, "%Y-%m-%d").date()
    except ValueError:
        start_date, end_date = datetime.now().date(), datetime.now().date()

    base_query = Order.objects.filter(created_at__date__gte=start_date, created_at__date__lte=end_date, is_void=False)
    if search_query:
        base_query = base_query.filter(id__icontains=search_query) | base_query.filter(customer__name__icontains=search_query)

    kpi_data = base_query.aggregate(total_rev=Sum('total_amount'), total_ord=Count('id'))
    total_revenue = float(kpi_data['total_rev'] or 0.00)
    total_orders = kpi_data['total_ord'] or 0
    express_orders_count = base_query.filter(items__is_express=True).distinct().count()
    total_items = OrderItem.objects.filter(order__in=base_query).aggregate(Sum('quantity'))['quantity__sum'] or 0
    days_diff = (end_date - start_date).days + 1
    daily_average = total_revenue / days_diff if days_diff > 0 else total_revenue

    # تجهيز ملف الإكسل (Excel)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Orders_Report_{start_date}_to_{end_date}.xlsx"'
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Orders Report"

    # تنسيقات الألوان والخطوط (الاحترافية)
    title_font = Font(bold=True, size=16, color="007EC5")
    bold_font = Font(bold=True)
    header_font = Font(bold=True, color="FFFFFF")
    blue_fill = PatternFill("solid", fgColor="007EC5")
    dark_fill = PatternFill("solid", fgColor="475569")
    center_align = Alignment(horizontal="center", vertical="center")

    # 1. الترويسة العلوية
    ws.append(['AL MARONA LAUNDRY - ORDERS REPORT'])
    ws['A1'].font = title_font
    ws.append([f'Date Range: {start_date} to {end_date}'])
    ws.append(['Generated on:', datetime.now().strftime('%Y-%m-%d %H:%M')])
    ws.append([])

    # 2. الملخص المالي (مع الألوان)
    ws.append(['--- SUMMARY OVERVIEW ---'])
    ws.cell(row=ws.max_row, column=1).font = bold_font
    
    summary_headers = ['Total Revenue (QAR)', 'Total Orders', 'Daily Average (QAR)', 'Express Orders', 'Total Items Processed']
    ws.append(summary_headers)
    for col in range(1, 6):
        cell = ws.cell(row=ws.max_row, column=col)
        cell.font = header_font
        cell.fill = blue_fill
        cell.alignment = center_align

    ws.append([round(total_revenue, 2), total_orders, round(daily_average, 2), express_orders_count, total_items])
    for col in range(1, 6):
        ws.cell(row=ws.max_row, column=col).alignment = center_align
        ws.cell(row=ws.max_row, column=col).font = bold_font
        
    ws.append([])
    ws.append([])

    # 3. تفاصيل الطلبات (الجدول الرئيسي)
    ws.append(['--- ORDER DETAILS ---'])
    ws.cell(row=ws.max_row, column=1).font = bold_font

    details_headers = ['Date', 'Order #', 'Customer', 'Items Qty', 'Amount (QAR)', 'Status']
    ws.append(details_headers)
    for col in range(1, 7):
        cell = ws.cell(row=ws.max_row, column=col)
        cell.font = header_font
        cell.fill = dark_fill

    for order in base_query.order_by('-created_at'):
        items_qty = order.items.aggregate(Sum('quantity'))['quantity__sum'] or 0
        customer_name = order.customer.name if order.customer else 'Cash Customer'
        ws.append([
            order.created_at.strftime('%Y-%m-%d %H:%M'),
            f"ORD-{order.id:04d}", 
            customer_name, 
            items_qty, 
            float(order.total_amount), 
            order.get_status_display()
        ])

    # 4. تعديل مقاسات الأعمدة تلقائياً عشان تطلع مرتبة
    column_widths = {'A': 20, 'B': 15, 'C': 30, 'D': 15, 'E': 18, 'F': 18}
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # حفظ وإرسال الملف
    wb.save(response)
    return response

# ==========================================
# 💰 قسم تقرير الإيرادات (Revenue Report) 💰
# ==========================================

def revenue_report_page(request):
    """ فتح صفحة تقرير الإيرادات """
    return render(request, 'revenue_report.html')

def api_revenue_report_data(request):
    """ API لجلب بيانات الإيرادات مجمعة (حسب اليوم والعميل) مع بحث ذكي """
    start_date_str = request.GET.get('start', '2000-01-01')
    end_date_str = request.GET.get('end', '2100-01-01')
    method_filter = request.GET.get('method', 'all').upper()
    search_query = request.GET.get('search', '').strip().lower()

    try:
        start_date = datetime.strptime(start_date_str, "%Y-%m-%d").date()
        end_date = datetime.strptime(end_date_str, "%Y-%m-%d").date()
    except ValueError:
        return JsonResponse({'error': 'Invalid date format'}, status=400)

    payments = Payment.objects.filter(
        created_at__date__gte=start_date,
        created_at__date__lte=end_date
    ).select_related('order', 'order__customer')

    if method_filter != 'ALL':
        db_method = 'CARD' if method_filter == 'BANK' else method_filter
        payments = payments.filter(method=db_method)

    if search_query:
        q_objects = Q(order__customer__name__icontains=search_query) | Q(order__customer__phone__icontains=search_query)
        extracted_num = ''.join(filter(str.isdigit, search_query))
        if extracted_num:
            num = int(extracted_num)
            q_objects |= Q(order__id=num)
            q_objects |= Q(id=num)
        payments = payments.filter(q_objects)

    # 🌟 هنا سحر التجميع: نجمع حسب (اليوم + العميل + طريقة الدفع) 🌟
    receipts_dict = {}
    for p in payments.order_by('created_at'):
        date_key = p.created_at.strftime('%Y-%m-%d')
        customer_id = p.order.customer.id if p.order.customer else 'walk_in_cash'
        key = f"{date_key}_{customer_id}_{p.method}"

        if key not in receipts_dict:
            c_name = p.order.customer.name if p.order.customer else 'Cash Customer'
            c_phone = p.order.customer.phone if p.order.customer else '---'
            c_file = p.order.customer.id if p.order.customer else '---'
            
            receipts_dict[key] = {
                'sort_date': p.created_at,
                'date': p.created_at.strftime('%d/%m/%Y'),
                'receipt_no': f"RCP-{p.id:06d}",
                'orders': [],
                'customer': c_name,
                'customer_file': c_file,
                'customer_phone': c_phone,
                'method': 'BANK' if p.method == 'CARD' else p.method,
                'amount': 0.0,
                'commission': 0.00
            }
        
        # 🌟 جعلنا الرقم يطابق النظام الأصلي OR-000000 🌟
        order_display = f"OR-{p.order.id:06d}"
        if not any(o['display'] == order_display for o in receipts_dict[key]['orders']):
            receipts_dict[key]['orders'].append({
                'id': p.order.id, 
                'display': order_display,
                'amount': float(p.amount)
            })
            
        receipts_dict[key]['amount'] += float(p.amount)

    receipts_list = sorted(list(receipts_dict.values()), key=lambda x: x['sort_date'], reverse=True)
    total_collections = sum(r['amount'] for r in receipts_list)
    cash_receipts = [r for r in receipts_list if r['method'] == 'CASH']
    bank_receipts = [r for r in receipts_list if r['method'] == 'BANK']

    return JsonResponse({
        'kpis': {
            'total_collections': round(total_collections, 2),
            'cash_total': round(sum(r['amount'] for r in cash_receipts), 2),
            'cash_count': len(cash_receipts),
            'bank_total': round(sum(r['amount'] for r in bank_receipts), 2),
            'bank_count': len(bank_receipts),
            'commission_total': 0.00
        },
        'table': receipts_list
    })

def print_grouped_receipt(request):
    """ طباعة سند القبض المجمع للرول الحراري (Actions) """
    order_ids = request.GET.get('orders', '').split(',')
    method = request.GET.get('method', 'CASH')
    receipt_no = request.GET.get('receipt_no', 'RCP-000000')
    
    payments = Payment.objects.filter(order__id__in=order_ids).select_related('order', 'order__customer')
    orders_data = []
    total_amount = 0
    customer = None
    
    for idx, p in enumerate(payments, 1):
        orders_data.append({
            'sn': idx,
            'order_no': f"OR-{p.order.id:06d}",
            'amount': p.amount
        })
        total_amount += p.amount
        if not customer and p.order.customer:
            customer = p.order.customer

    context = {
        'receipt_no': receipt_no,
        'date': timezone.now().strftime('%B %dth, %Y'),
        'time': timezone.now().strftime('%I:%M %p'),
        'method': method,
        'total_amount': total_amount,
        'orders': orders_data,
        'customer_name': customer.name if customer else 'Cash Customer',
        'customer_code': customer.id if customer else '---',
    }
    return render(request, 'payment_receipt_thermal.html', context)

def export_revenue_excel(request):
    """ تصدير الإكسل بنفس نظام التجميع الجديد (حسب اليوم والعميل) """
    start_date_str = request.GET.get('start', '2000-01-01')
    end_date_str = request.GET.get('end', '2100-01-01')
    method_filter = request.GET.get('method', 'all').upper()
    search_query = request.GET.get('search', '').strip().lower()

    try:
        start_date = datetime.strptime(start_date_str, "%Y-%m-%d").date()
        end_date = datetime.strptime(end_date_str, "%Y-%m-%d").date()
    except ValueError:
        start_date, end_date = datetime.now().date(), datetime.now().date()

    payments = Payment.objects.filter(created_at__date__gte=start_date, created_at__date__lte=end_date).select_related('order', 'order__customer')
    if method_filter != 'ALL':
        db_method = 'CARD' if method_filter == 'BANK' else method_filter
        payments = payments.filter(method=db_method)

    if search_query:
        q_objects = Q(order__customer__name__icontains=search_query) | Q(order__customer__phone__icontains=search_query)
        extracted_num = ''.join(filter(str.isdigit, search_query))
        if extracted_num:
            num = int(extracted_num)
            q_objects |= Q(order__id=num)
            q_objects |= Q(id=num)
        payments = payments.filter(q_objects)

    receipts_dict = {}
    for p in payments.order_by('created_at'):
        date_key = p.created_at.strftime('%Y-%m-%d')
        customer_id = p.order.customer.id if p.order.customer else 'walk_in_cash'
        key = f"{date_key}_{customer_id}_{p.method}"

        if key not in receipts_dict:
            receipts_dict[key] = {
                'sort_date': p.created_at,
                'date': p.created_at.strftime('%d/%m/%Y'),
                'receipt_no': f"RCP-{p.id:06d}",
                'orders': [],
                'customer': p.order.customer.name if p.order.customer else 'Cash Customer',
                'method': 'BANK' if p.method == 'CARD' else p.method,
                'amount': 0.0
            }
            
        order_display = f"OR-{p.order.id:06d}"
        if order_display not in receipts_dict[key]['orders']:
            receipts_dict[key]['orders'].append(order_display)
            
        receipts_dict[key]['amount'] += float(p.amount)

    receipts_list = sorted(list(receipts_dict.values()), key=lambda x: x['sort_date'], reverse=True)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Revenue_Report_{start_date}_to_{end_date}.xlsx"'
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Collections Report"

    title_font = Font(bold=True, size=16, color="007EC5")
    header_font = Font(bold=True, color="FFFFFF")
    blue_fill = PatternFill("solid", fgColor="0f172a")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    ws.append(['AL MARONA LAUNDRY - COLLECTIONS REPORT'])
    ws['A1'].font = title_font
    ws.append([f'Date Range: {start_date} to {end_date}'])
    ws.append(['Generated on:', datetime.now().strftime('%Y-%m-%d %H:%M')])
    ws.append([])

    headers = ['Date', 'Receipt Number', 'Order Numbers', 'Customer', 'Payment Method', 'Bank Account', 'Amount (QAR)', 'Commission']
    ws.append(headers)
    
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=ws.max_row, column=col)
        cell.font = header_font
        cell.fill = blue_fill
        cell.border = thin_border

    for r in receipts_list:
        orders_str = ", ".join(r['orders'])
        row_data = [
            r['date'], r['receipt_no'], orders_str, r['customer'], 
            r['method'], '-', r['amount'], 0.00
        ]
        ws.append(row_data)
        for col in range(1, len(row_data) + 1):
            ws.cell(row=ws.max_row, column=col).border = thin_border

    column_widths = {'A': 15, 'B': 20, 'C': 40, 'D': 25, 'E': 20, 'F': 15, 'G': 15, 'H': 15}
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    wb.save(response)
    return response

# ==========================================
# ⚖️ قسم الميزانية العمومية (Balance Sheet) ⚖️
# ==========================================

def balance_sheet_page(request):
    """ فتح صفحة الميزانية العمومية """
    return render(request, 'laundry_balance_sheet.html')

def api_balance_sheet_data(request):
    """ API لحساب الميزانية العمومية الموزونة 100% """
    as_of_str = request.GET.get('date', timezone.now().strftime('%Y-%m-%d'))
    try:
        as_of_date = datetime.strptime(as_of_str, "%Y-%m-%d").date()
    except ValueError:
        as_of_date = timezone.now().date()

    # 1. الذمم المدينة (Sundry Debtors)
    orders = Order.objects.filter(created_at__date__lte=as_of_date, is_void=False)
    debtors_dict = {}
    for order in orders:
        if order.customer:
            paid = sum(p.amount for p in order.payments.filter(created_at__date__lte=as_of_date))
            owed = float(order.total_amount) - float(paid)
            if owed > 0.01:
                c_id = order.customer.id
                if c_id not in debtors_dict:
                    debtors_dict[c_id] = {"code": f"ASS-{c_id:04d}", "name": order.customer.name or order.customer.phone or f"Customer {c_id}", "amount": 0.0}
                debtors_dict[c_id]["amount"] += owed

    debtors_list = sorted(list(debtors_dict.values()), key=lambda x: x["amount"], reverse=True)
    sundry_debtors_total = sum(d["amount"] for d in debtors_list)

    # 2. الكاش والبنك (Cash in Hand & Bank Balances)
    payments = Payment.objects.filter(created_at__date__lte=as_of_date)
    cash_inflows = sum(float(p.amount) for p in payments if p.method == 'CASH')
    bank_inflows = sum(float(p.amount) for p in payments if p.method in ['CARD', 'BANK'])

    expenses = Expense.objects.filter(date__lte=as_of_date)
    cash_outflows = sum(float(e.paid_amount) for e in expenses if e.payment_method == 'CASH')
    bank_outflows = sum(float(e.paid_amount) for e in expenses if e.payment_method in ['CARD', 'BANK'])

    cash_in_hand = cash_inflows - cash_outflows
    bank_balance = bank_inflows - bank_outflows

    total_assets = sundry_debtors_total + cash_in_hand + bank_balance

    # 3. الالتزامات (Liabilities)
    liabilities_total = sum((float(e.total_amount) - float(e.paid_amount)) for e in expenses if e.total_amount > e.paid_amount)

    # 4. حقوق الملكية (Equity)
    total_revenue = sum(float(o.total_amount) for o in orders)
    total_expense_incurred = sum(float(e.total_amount) for e in expenses)
    retained_earnings = total_revenue - total_expense_incurred

    return JsonResponse({
        "as_of_date": as_of_date.strftime("%B %d, %Y"),
        "liabilities": {
            "total": round(liabilities_total, 2)
        },
        "equity": {
            "retained_earnings": round(retained_earnings, 2),
            "total": round(retained_earnings, 2)
        },
        "assets": {
            "cash_in_hand": round(cash_in_hand, 2),
            "bank_balance": round(bank_balance, 2),
            "sundry_debtors_total": round(sundry_debtors_total, 2),
            "total_assets": round(total_assets, 2),
            "debtors": debtors_list
        }
    })

def export_balance_sheet_excel(request):
    """ تصدير الميزانية الموزونة لإكسل """
    as_of_str = request.GET.get('date', timezone.now().strftime('%Y-%m-%d'))
    try:
        as_of_date = datetime.strptime(as_of_str, "%Y-%m-%d").date()
    except ValueError:
        as_of_date = timezone.now().date()

    orders = Order.objects.filter(created_at__date__lte=as_of_date, is_void=False)
    debtors_dict = {}
    for order in orders:
        if order.customer:
            paid = sum(p.amount for p in order.payments.filter(created_at__date__lte=as_of_date))
            owed = float(order.total_amount) - float(paid)
            if owed > 0.01:
                c_id = order.customer.id
                if c_id not in debtors_dict:
                    debtors_dict[c_id] = {"code": f"ASS-{c_id:04d}", "name": order.customer.name or order.customer.phone, "amount": 0.0}
                debtors_dict[c_id]["amount"] += owed

    debtors_list = sorted(list(debtors_dict.values()), key=lambda x: x["amount"], reverse=True)
    sundry_debtors_total = sum(d["amount"] for d in debtors_list)

    payments = Payment.objects.filter(created_at__date__lte=as_of_date)
    cash_inflows = sum(float(p.amount) for p in payments if p.method == 'CASH')
    bank_inflows = sum(float(p.amount) for p in payments if p.method in ['CARD', 'BANK'])

    expenses = Expense.objects.filter(date__lte=as_of_date)
    cash_outflows = sum(float(e.paid_amount) for e in expenses if e.payment_method == 'CASH')
    bank_outflows = sum(float(e.paid_amount) for e in expenses if e.payment_method in ['CARD', 'BANK'])

    cash_in_hand = cash_inflows - cash_outflows
    bank_balance = bank_inflows - bank_outflows
    total_assets = sundry_debtors_total + cash_in_hand + bank_balance

    liabilities_total = sum((float(e.total_amount) - float(e.paid_amount)) for e in expenses if e.total_amount > e.paid_amount)
    total_revenue = sum(float(o.total_amount) for o in orders)
    total_expense_incurred = sum(float(e.total_amount) for e in expenses)
    retained_earnings = total_revenue - total_expense_incurred

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Balance_Sheet_{as_of_date}.xlsx"'
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Balance Sheet"

    title_font = Font(bold=True, size=16, color="007EC5")
    header_font = Font(bold=True, color="FFFFFF")
    blue_fill = PatternFill("solid", fgColor="0f172a")
    light_gray = PatternFill("solid", fgColor="f1f5f9")

    ws.append(['AL MARONA LAUNDRY - BALANCE SHEET (الميزانية العمومية)'])
    ws['A1'].font = title_font
    ws.append([f'As of Date: {as_of_date.strftime("%B %d, %Y")}'])
    ws.append([])

    ws.append(['Liabilities (الالتزامات)', 'Amount (QAR)'])
    for col in range(1, 3): ws.cell(row=ws.max_row, column=col).font = header_font; ws.cell(row=ws.max_row, column=col).fill = blue_fill
    ws.append(['Total Liabilities', round(liabilities_total, 2)])
    ws.append([])

    ws.append(['Equity (حقوق الملكية)', 'Amount (QAR)'])
    for col in range(1, 3): ws.cell(row=ws.max_row, column=col).font = header_font; ws.cell(row=ws.max_row, column=col).fill = blue_fill
    ws.append(['Retained Earnings (الأرباح المحتجزة)', round(retained_earnings, 2)])
    ws.append(['Total Equity', round(retained_earnings, 2)])
    ws.append([])

    ws.append(['Assets (الأصول)', 'Code', 'Amount (QAR)'])
    for col in range(1, 4): ws.cell(row=ws.max_row, column=col).font = header_font; ws.cell(row=ws.max_row, column=col).fill = blue_fill
    
    ws.append(['Cash in Hand (النقد في الصندوق)', '-', round(cash_in_hand, 2)])
    ws.append(['Bank Balances (حسابات البنك)', '-', round(bank_balance, 2)])
    ws.append(['Sundry Debtors (الذمم المدينة)', '-', round(sundry_debtors_total, 2)])
    
    for d in debtors_list:
        ws.append([d['name'], d['code'], round(d['amount'], 2)])
    
    ws.append(['Total Assets (إجمالي الأصول)', '', round(total_assets, 2)])
    for col in range(1, 4): ws.cell(row=ws.max_row, column=col).font = Font(bold=True); ws.cell(row=ws.max_row, column=col).fill = light_gray

    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20

    wb.save(response)
    return response

# ==========================================
# 💸 تقرير التحصيل / ديون العملاء (Outstanding Customers List) 💸
# ==========================================

def collection_report_page(request):
    """ فتح صفحة تقرير التحصيل (ديون العملاء) """
    return render(request, 'collection_report.html')

def api_collection_report_data(request):
    """ API لجلب العملاء المديونين """
    as_of_str = request.GET.get('date', timezone.now().strftime('%Y-%m-%d'))
    try:
        as_of_date = datetime.strptime(as_of_str, "%Y-%m-%d").date()
    except ValueError:
        as_of_date = timezone.now().date()

    orders = Order.objects.filter(created_at__date__lte=as_of_date, is_void=False)
    debtors_dict = {}

    for order in orders:
        if order.customer:
            paid = sum(float(p.amount or 0) for p in order.payments.filter(created_at__date__lte=as_of_date))
            owed = float(order.total_amount or 0) - paid

            if owed > 0.01:
                c_id = order.customer.id
                if c_id not in debtors_dict:
                    # التحقق إذا كان في حقل للعنوان، وإلا نخليه فاضي
                    address = getattr(order.customer, 'address', '')
                    
                    debtors_dict[c_id] = {
                        "name": order.customer.name or f"Customer {c_id}",
                        "file_no": str(c_id),
                        "phone": order.customer.phone or "",
                        "address": address or "",
                        "category": "Uncategorized",
                        "type": "CREDIT",
                        "amount": 0.0
                    }
                debtors_dict[c_id]["amount"] += owed

    # ترتيب العملاء أبجدياً
    debtors_list = sorted(list(debtors_dict.values()), key=lambda x: x["name"].lower())
    total_outstanding = sum(d["amount"] for d in debtors_list)

    return JsonResponse({
        "as_of_date": as_of_date.strftime("%d/%m/%Y"),
        "kpis": {
            "total_outstanding": round(total_outstanding, 2),
            "total_debtors": len(debtors_list)
        },
        "table": debtors_list
    })

def export_collection_excel(request):
    """ تصدير ديون العملاء لملف إكسل خرافي """
    as_of_str = request.GET.get('date', timezone.now().strftime('%Y-%m-%d'))
    try:
        as_of_date = datetime.strptime(as_of_str, "%Y-%m-%d").date()
    except ValueError:
        as_of_date = timezone.now().date()

    orders = Order.objects.filter(created_at__date__lte=as_of_date, is_void=False)
    debtors_dict = {}

    for order in orders:
        if order.customer:
            paid = sum(float(p.amount or 0) for p in order.payments.filter(created_at__date__lte=as_of_date))
            owed = float(order.total_amount or 0) - paid
            if owed > 0.01:
                c_id = order.customer.id
                if c_id not in debtors_dict:
                    address = getattr(order.customer, 'address', '')
                    debtors_dict[c_id] = {
                        "name": order.customer.name or f"Customer {c_id}",
                        "file_no": str(c_id),
                        "phone": order.customer.phone or "",
                        "address": address or "",
                        "category": "Uncategorized",
                        "type": "CREDIT",
                        "amount": 0.0
                    }
                debtors_dict[c_id]["amount"] += owed

    debtors_list = sorted(list(debtors_dict.values()), key=lambda x: x["name"].lower())
    total_outstanding = sum(d["amount"] for d in debtors_list)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Outstanding_Customers_{as_of_date}.xlsx"'
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Outstanding Customers List"

    title_font = Font(bold=True, size=18, color="B91C1C") # لون أحمر داكن للديون
    header_font = Font(bold=True, color="FFFFFF")
    dark_fill = PatternFill("solid", fgColor="1F2937")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    center_align = Alignment(horizontal="center", vertical="center")

    ws.append(['AL MARONA LAUNDRY - OUTSTANDING CUSTOMERS LIST'])
    ws['A1'].font = title_font
    ws.append([f'Generated on: {timezone.now().strftime("%d/%m/%Y %H:%M")}'])
    ws.append([f'Total Outstanding Amount: QAR {round(total_outstanding, 2)}'])
    ws['A3'].font = Font(bold=True, color="B91C1C", size=14)
    ws.append([])

    headers = ['Name', 'File Number', 'Phone Number', 'Address', 'Customer Category', 'Customer Type', 'Outstanding Amount (QAR)']
    ws.append(headers)
    
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=ws.max_row, column=col)
        cell.font = header_font
        cell.fill = dark_fill
        cell.border = thin_border
        cell.alignment = center_align

    for d in debtors_list:
        row_data = [d['name'], d['file_no'], d['phone'], d['address'], d['category'], d['type'], round(d['amount'], 2)]
        ws.append(row_data)
        for col in range(1, len(row_data) + 1):
            cell = ws.cell(row=ws.max_row, column=col)
            cell.border = thin_border
            if col == 7: # لون أحمر للمبلغ
                cell.font = Font(bold=True, color="B91C1C")

    # إجمالي في آخر الجدول
    ws.append(['', '', '', '', '', 'Total Outstanding:', round(total_outstanding, 2)])
    ws.cell(row=ws.max_row, column=6).font = Font(bold=True)
    ws.cell(row=ws.max_row, column=7).font = Font(bold=True, color="B91C1C")

    column_widths = {'A': 25, 'B': 15, 'C': 20, 'D': 25, 'E': 20, 'F': 15, 'G': 25}
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    wb.save(response)
    return response
# ==========================================
# 📊 كشف الحساب التفصيلي (Customer Ledger) 📊
# ==========================================

# ==========================================
# 📊 كشف الحساب التفصيلي (Customer Ledger) 📊
# ==========================================

def customer_ledger_page(request):
    """ فتح صفحة كشف الحساب المدمجة (معدلة وآمنة) """
    unique_customers = Order.objects.exclude(customer__isnull=True).values(
        'customer__id', 'customer__name', 'customer__phone'
    ).distinct()
    
    customers = []
    for c in unique_customers:
        customers.append({
            'id': c['customer__id'], 
            'name': c['customer__name'] or f"Customer {c['customer__id']}", 
            'phone': c['customer__phone'] or ''
        })
        
    customers = sorted(customers, key=lambda k: k['name'])
    return render(request, 'customer_ledger.html', {'customers': customers})

def api_ledger_report_data(request):
    """ API محمي ومصفح ضد أي بيانات ناقصة في قاعدة البيانات """
    try:
        customer_id = request.GET.get('customer_id', 'all')
        start_date_str = request.GET.get('start', (timezone.now() - timedelta(days=30)).strftime('%Y-%m-%d'))
        end_date_str = request.GET.get('end', timezone.now().strftime('%Y-%m-%d'))

        try:
            start_date = datetime.strptime(start_date_str, "%Y-%m-%d").date()
            end_date = datetime.strptime(end_date_str, "%Y-%m-%d").date()
        except ValueError:
            start_date = (timezone.now() - timedelta(days=30)).date()
            end_date = timezone.now().date()

        account_name = "All Customers | كل العملاء"
        account_code = "GEN-000"
        file_no = "-"
        phone = "-"
        address = "N/A"
        
        orders = Order.objects.filter(created_at__date__gte=start_date, created_at__date__lte=end_date, is_void=False)
        payments = Payment.objects.filter(created_at__date__gte=start_date, created_at__date__lte=end_date)

        if customer_id and customer_id != 'all':
            orders = orders.filter(customer_id=customer_id)
            payments = payments.filter(order__customer_id=customer_id)
            
            c_info = Order.objects.filter(customer_id=customer_id).values(
                'customer__id', 'customer__name', 'customer__phone'
            ).first()
            
            if c_info:
                account_name = c_info.get('customer__name') or "Unknown"
                c_id = c_info.get('customer__id') or 0
                try:
                    account_code = f"ASS-{int(c_id):04d}"
                except:
                    account_code = f"ASS-{c_id}"
                file_no = str(c_id)
                phone = c_info.get('customer__phone') or "-"

        transactions = []
        for order in orders:
            transactions.append({
                "date": order.created_at.strftime('%Y-%m-%d'),
                "date_obj": order.created_at,
                "ref": f"INV-{order.id:06d}",
                "particulars": "Laundry Services",
                "description": f"Order {order.id}",
                "debit": float(order.total_amount or 0.0),
                "credit": 0.0
            })

        for pay in payments:
            order_ref = str(pay.order.id) if getattr(pay, 'order', None) else 'Unknown'
            transactions.append({
                "date": pay.created_at.strftime('%Y-%m-%d'),
                "date_obj": pay.created_at,
                "ref": f"RCP-{pay.id:06d}",
                "particulars": f"{pay.method} Payment",
                "description": f"Payment for Order {order_ref}",
                "debit": 0.0,
                "credit": float(pay.amount or 0.0)
            })

        transactions.sort(key=lambda x: x["date_obj"])

        balance = 0.0
        total_debits = 0.0
        total_credits = 0.0
        chart_labels = []
        chart_data = []

        for t in transactions:
            total_debits += t["debit"]
            total_credits += t["credit"]
            balance += (t["debit"] - t["credit"])
            t["balance"] = balance
            chart_labels.append(t["date"])
            chart_data.append(balance)
            del t["date_obj"]

        print_items = []
        total_print_items = 0
        total_print_amount = 0.0
        
        pending_orders = Order.objects.filter(customer_id=customer_id, is_void=False) if customer_id != 'all' else []
        for po in pending_orders:
            paid = sum(float(p.amount or 0.0) for p in po.payments.all())
            owed = float(po.total_amount or 0.0) - paid
            
            if owed > 0.01:
                total_print_amount += float(po.total_amount or 0.0)
                is_first_item = True
                
                for item in po.items.all():
                    qty = float(getattr(item, 'quantity', 1))
                    price = float(getattr(item, 'price', 0.0))
                    total_print_items += qty
                    
                    # استخراج اسم الخدمة بطريقة آمنة جداً (تمنع توقف النظام)
                    item_name = "Service Item"
                    service_cat = "Washing"
                    
                    if hasattr(item, 'service_item') and item.service_item:
                        item_name = getattr(item.service_item, 'name', 'Service Item')
                        service_cat = getattr(item.service_item, 'category', 'Washing')
                    elif hasattr(item, 'service') and item.service:
                        item_name = getattr(item.service, 'name', 'Service Item')
                        service_cat = getattr(item.service, 'category', 'Washing')
                    elif hasattr(item, 'name'):
                        item_name = getattr(item, 'name', 'Service Item')

                    print_items.append({
                        "order_no": f"OR-{po.id:06d}",
                        "date": po.created_at.strftime('%B %d, %Y'),
                        "item_name": str(item_name),
                        "service": str(service_cat),
                        "qty": int(qty),
                        "unit_price": price,
                        "item_total": qty * price,
                        "order_total": float(po.total_amount or 0.0) if is_first_item else "", 
                        "discount": getattr(po, 'discount', "-") if is_first_item else ""
                    })
                    is_first_item = False

        return JsonResponse({
            "account": {
                "code": account_code,
                "name": account_name,
                "group": "Sundry Debtors" if customer_id != 'all' else "All Accounts",
                "file_no": file_no,
                "phone": phone,
                "address": address,
                "opening_balance": 0.00,
                "current_balance": round(balance, 2)
            },
            "summary": {
                "start": start_date_str,
                "end": end_date_str,
                "total_debits": round(total_debits, 2),
                "total_credits": round(total_credits, 2)
            },
            "transactions": transactions,
            "chart": {
                "labels": chart_labels,
                "data": chart_data
            },
            "print_invoice": {
                "items": print_items,
                "total_qty": int(total_print_items),
                "grand_total": round(total_print_amount, 2)
            }
        })
    except Exception as e:
        # إرجاع الخطأ للواجهة عشان ما يعلق النظام
        return JsonResponse({"error": str(e)}, status=500)


def export_ledger_excel(request):
    """ تصدير كشف الحساب بملف إكسل احترافي وآمن """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from datetime import datetime, timedelta
    from django.utils import timezone

    customer_id = request.GET.get('customer_id', 'all')
    start_date_str = request.GET.get('start', (timezone.now() - timedelta(days=30)).strftime('%Y-%m-%d'))
    end_date_str = request.GET.get('end', timezone.now().strftime('%Y-%m-%d'))

    try:
        start_date = datetime.strptime(start_date_str, "%Y-%m-%d").date()
        end_date = datetime.strptime(end_date_str, "%Y-%m-%d").date()
    except ValueError:
        start_date = (timezone.now() - timedelta(days=30)).date()
        end_date = timezone.now().date()

    orders = Order.objects.filter(created_at__date__gte=start_date, created_at__date__lte=end_date, is_void=False)
    payments = Payment.objects.filter(created_at__date__gte=start_date, created_at__date__lte=end_date)
    account_name = "All Customers"

    if customer_id != 'all' and customer_id:
        orders = orders.filter(customer_id=customer_id)
        payments = payments.filter(order__customer_id=customer_id)
        c_info = Order.objects.filter(customer_id=customer_id).values('customer__name').first()
        if c_info:
            account_name = c_info.get('customer__name') or "Unknown"

    transactions = []
    for order in orders:
        transactions.append({
            "date": order.created_at.strftime('%Y-%m-%d'),
            "date_obj": order.created_at,
            "ref": f"INV-{order.id:06d}",
            "particulars": "Laundry Services",
            "description": f"Order {order.id}",
            "debit": float(order.total_amount or 0.0),
            "credit": 0.0
        })

    for pay in payments:
        order_ref = str(pay.order.id) if getattr(pay, 'order', None) else 'Unknown'
        transactions.append({
            "date": pay.created_at.strftime('%Y-%m-%d'),
            "date_obj": pay.created_at,
            "ref": f"RCP-{pay.id:06d}",
            "particulars": f"{pay.method} Payment",
            "description": f"Payment for Order {order_ref}",
            "debit": 0.0,
            "credit": float(pay.amount or 0.0)
        })

    transactions.sort(key=lambda x: x["date_obj"])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Ledger_Report_{start_date}_to_{end_date}.xlsx"'
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ledger Report"

    title_font = Font(bold=True, size=16, color="007EC5")
    header_font = Font(bold=True, color="FFFFFF")
    blue_fill = PatternFill("solid", fgColor="0f172a")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    ws.append(['AL MARONA LAUNDRY - LEDGER REPORT (كشف حساب)'])
    ws['A1'].font = title_font
    ws.append([f'Account: {account_name}'])
    ws.append([f'Period: {start_date} to {end_date}'])
    ws.append([])

    headers = ['Date', 'Reference', 'Particulars', 'Description', 'Debit (+)', 'Credit (-)', 'Balance']
    ws.append(headers)
    
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=ws.max_row, column=col)
        cell.font = header_font
        cell.fill = blue_fill
        cell.border = thin_border

    balance = 0.0
    for t in transactions:
        balance += (t["debit"] - t["credit"])
        row_data = [t["date"], t["ref"], t["particulars"], t["description"], t["debit"] or "-", t["credit"] or "-", balance]
        ws.append(row_data)
        for col in range(1, len(row_data) + 1):
            ws.cell(row=ws.max_row, column=col).border = thin_border

    column_widths = {'A': 15, 'B': 15, 'C': 25, 'D': 30, 'E': 15, 'F': 15, 'G': 15}
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    wb.save(response)
    return response

# ==========================================
# 📈 تقرير ربحية المنتجات (Business Performance) 📈
# ==========================================

def profitability_report_page(request):
    """ فتح صفحة ربحية المنتجات (مدمجة اللغات) """
    return render(request, 'product_profitability.html')

def api_profitability_report_data(request):
    """ API لحساب إيرادات وتكاليف المنتجات وصافي الربح """
    try:
        start_date_str = request.GET.get('start', (timezone.now() - timedelta(days=30)).strftime('%Y-%m-%d'))
        end_date_str = request.GET.get('end', timezone.now().strftime('%Y-%m-%d'))
        search_query = request.GET.get('search', '').lower().strip()
        cost_filter = request.GET.get('cost_filter', 'all')

        try:
            start_date = datetime.strptime(start_date_str, "%Y-%m-%d").date()
            end_date = datetime.strptime(end_date_str, "%Y-%m-%d").date()
        except ValueError:
            start_date = (timezone.now() - timedelta(days=30)).date()
            end_date = timezone.now().date()

        # 1. جلب الطلبات والمدفوعات
        orders = Order.objects.filter(created_at__date__gte=start_date, created_at__date__lte=end_date, is_void=False)
        payments = Payment.objects.filter(created_at__date__gte=start_date, created_at__date__lte=end_date)

        total_earnings = sum(float(o.total_amount or 0) for o in orders)
        receipts_count = orders.count()

        cash_receipts = payments.filter(method='CASH')
        bank_receipts = payments.exclude(method='CASH')
        
        cash_total = sum(float(p.amount or 0) for p in cash_receipts)
        bank_total = sum(float(p.amount or 0) for p in bank_receipts)
        
        cash_count = cash_receipts.count()
        bank_count = bank_receipts.count()

        # 2. حساب تفاصيل المنتجات والتكاليف
        products_dict = {}
        total_usage_cost = 0.0
        total_products_used = 0

        # جلب كل العناصر المباعة في هذه الطلبات
        order_items = OrderItem.objects.filter(order__in=orders).select_related('order')
        
        for item in order_items:
            qty = float(getattr(item, 'quantity', 1))
            # استخراج اسم المنتج والتكلفة بأمان
            prod_name = "Unknown Product"
            category = "General"
            unit_cost = 0.0 # افتراضياً صفر إذا لم تكن هناك تكلفة مسجلة
            
            if hasattr(item, 'service_item') and item.service_item:
                prod_name = getattr(item.service_item, 'name', 'Unknown Product')
                category = getattr(item.service_item, 'category', 'General')
                unit_cost = float(getattr(item.service_item, 'cost', 0.0))
            elif hasattr(item, 'name'):
                prod_name = getattr(item, 'name', 'Unknown Product')

            total_item_cost = qty * unit_cost
            
            if prod_name not in products_dict:
                products_dict[prod_name] = {
                    'name': prod_name,
                    'category': category,
                    'qty': 0,
                    'unit_cost': unit_cost,
                    'total_cost': 0.0
                }
            
            products_dict[prod_name]['qty'] += qty
            products_dict[prod_name]['total_cost'] += total_item_cost
            total_usage_cost += total_item_cost
            total_products_used += qty

        # 3. الفلترة (بحث وتكلفة)
        products_list = list(products_dict.values())
        
        if search_query:
            products_list = [p for p in products_list if search_query in p['name'].lower() or search_query in p['category'].lower()]
            
        if cost_filter == 'high':
            products_list = [p for p in products_list if p['total_cost'] > 100]
        elif cost_filter == 'low':
            products_list = [p for p in products_list if p['total_cost'] <= 100]

        # 4. حساب المؤشرات (KPIs)
        net_profit = total_earnings - total_usage_cost
        profit_margin = (net_profit / total_earnings * 100) if total_earnings > 0 else 0.0
        efficiency_ratio = (total_earnings / total_usage_cost) if total_usage_cost > 0 else 0.0
        cost_percentage = (total_usage_cost / total_earnings * 100) if total_earnings > 0 else 0.0

        return JsonResponse({
            "summary": {
                "start": start_date_str,
                "end": end_date_str,
                "total_earnings": round(total_earnings, 2),
                "receipts_count": receipts_count,
                "cash_total": round(cash_total, 2),
                "cash_count": cash_count,
                "bank_total": round(bank_total, 2),
                "bank_count": bank_count,
                "usage_costs": round(total_usage_cost, 2),
                "products_used": int(total_products_used),
                "categories_count": len(set(p['category'] for p in products_list)),
                "net_profit": round(net_profit, 2),
                "profit_margin": round(profit_margin, 2),
                "efficiency_ratio": round(efficiency_ratio, 2),
                "cost_percentage": round(cost_percentage, 2)
            },
            "products": sorted(products_list, key=lambda x: x['total_cost'], reverse=True)
        })
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


def export_profitability_excel(request):
    """ تصدير تقرير الربحية بملف إكسل إبداعي """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    # استدعاء نفس بيانات الـ API داخلياً لتجنب تكرار الكود
    data_response = api_profitability_report_data(request)
    if data_response.status_code != 200:
        return HttpResponse("Error generating report", status=500)
    
    import json
    data = json.loads(data_response.content)
    summary = data['summary']
    products = data['products']

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Business_Performance_{summary["start"]}_to_{summary["end"]}.xlsx"'
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Performance Report"

    # تنسيقات
    title_font = Font(bold=True, size=18, color="0f172a")
    header_font = Font(bold=True, color="FFFFFF")
    blue_fill = PatternFill("solid", fgColor="0ea5e9")
    green_fill = PatternFill("solid", fgColor="10b981")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # الترويسة
    ws.append(['BUSINESS PERFORMANCE REPORT (تقرير أداء العمل)'])
    ws['A1'].font = title_font
    ws.append([f'Period: {summary["start"]} to {summary["end"]}'])
    ws.append([])

    # ملخص الأداء
    ws.append(['--- PERFORMANCE SUMMARY ---'])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True, color="0ea5e9")
    
    ws.append(['Total Earnings (إجمالي الإيرادات)', round(summary['total_earnings'], 2), 'Usage Costs (تكاليف الاستخدام)', round(summary['usage_costs'], 2)])
    ws.append(['Net Profit (صافي الربح)', round(summary['net_profit'], 2), 'Profit Margin (هامش الربح)', f"{summary['profit_margin']}%"])
    ws.append([])

    # جدول المنتجات
    headers = ['Product (المنتج)', 'Category (الفئة)', 'Quantity Used (الكمية المستخدمة)', 'Unit Cost (تكلفة الوحدة)', 'Total Cost (إجمالي التكلفة)']
    ws.append(headers)
    
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=ws.max_row, column=col)
        cell.font = header_font
        cell.fill = blue_fill
        cell.border = thin_border

    for p in products:
        row_data = [p['name'], p['category'], p['qty'], round(p['unit_cost'], 2), round(p['total_cost'], 2)]
        ws.append(row_data)
        for col in range(1, len(row_data) + 1):
            ws.cell(row=ws.max_row, column=col).border = thin_border

    # ضبط المقاسات
    column_widths = {'A': 30, 'B': 20, 'C': 25, 'D': 20, 'E': 20}
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    wb.save(response)
    return response    


# ==========================================
# 📈 تقرير الربح والخسارة (Profit & Loss) 📈
# ==========================================

def profit_loss_page(request):
    """ فتح صفحة الربح والخسارة """
    return render(request, 'profit_loss.html')

def api_profit_loss_data(request):
    """ API لحساب قائمة الدخل (حساب المتاجرة والأرباح والخسائر) """
    try:
        start_date_str = request.GET.get('start', (timezone.now() - timedelta(days=30)).strftime('%Y-%m-%d'))
        end_date_str = request.GET.get('end', timezone.now().strftime('%Y-%m-%d'))

        try:
            start_date = datetime.strptime(start_date_str, "%Y-%m-%d").date()
            end_date = datetime.strptime(end_date_str, "%Y-%m-%d").date()
        except ValueError:
            start_date = (timezone.now() - timedelta(days=30)).date()
            end_date = timezone.now().date()

        # 1. المبيعات (Sales)
        orders = Order.objects.filter(created_at__date__gte=start_date, created_at__date__lte=end_date, is_void=False)
        sales_total = sum(float(o.total_amount or 0) for o in orders)

        # 2. المصروفات (Expenses)
        expenses = Expense.objects.filter(date__gte=start_date, date__lte=end_date)
        
        indirect_expenses_list = []
        total_indirect_expenses = 0.0
        
        # تجميع المصروفات حسب الفئة
        from django.db.models import Sum
        expense_summary = expenses.values('category__name').annotate(total=Sum('total_amount'))
        for e in expense_summary:
            cat_name = e['category__name'] or 'General Expenses'
            amt = float(e['total'] or 0)
            indirect_expenses_list.append({"name": f"EXP - {cat_name}", "amount": amt})
            total_indirect_expenses += amt

        # --- حساب المتاجرة (Trading Account) ---
        opening_stock = 0.0
        purchases = 0.0
        direct_expenses = 0.0
        closing_stock = 0.0
        direct_income = 0.0

        total_trading_dr = opening_stock + purchases + direct_expenses
        total_trading_cr = sales_total + closing_stock + direct_income
        gross_profit = total_trading_cr - total_trading_dr
        trading_grand_total = max(total_trading_dr, total_trading_cr)

        # --- حساب الأرباح والخسائر (P&L Account) ---
        indirect_income = 0.0
        total_pl_cr = gross_profit + indirect_income
        total_pl_dr = total_indirect_expenses
        net_profit = total_pl_cr - total_pl_dr
        pl_grand_total = max(total_pl_dr, total_pl_cr)

        return JsonResponse({
            "period": f"{start_date.strftime('%d %B %Y')} to {end_date.strftime('%d %B %Y')}",
            "start": start_date_str,
            "end": end_date_str,
            "trading": {
                "opening_stock": round(opening_stock, 2),
                "purchases": round(purchases, 2),
                "direct_expenses": round(direct_expenses, 2),
                "sales": round(sales_total, 2),
                "closing_stock": round(closing_stock, 2),
                "direct_income": round(direct_income, 2),
                "gross_profit": round(gross_profit, 2),
                "grand_total": round(trading_grand_total, 2)
            },
            "pl": {
                "indirect_expenses": indirect_expenses_list,
                "total_indirect_expenses": round(total_indirect_expenses, 2),
                "gross_profit_bf": round(gross_profit, 2),
                "indirect_income": round(indirect_income, 2),
                "net_profit": round(net_profit, 2),
                "grand_total": round(pl_grand_total, 2)
            }
        })
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


def export_profit_loss_excel(request):
    """ تصدير الربح والخسارة لملف إكسل إبداعي ومطابق للشكل المحاسبي """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    import json

    data_response = api_profit_loss_data(request)
    if data_response.status_code != 200:
        return HttpResponse("Error generating report", status=500)
    
    data = json.loads(data_response.content)
    trading = data['trading']
    pl = data['pl']

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Profit_Loss_{data["start"]}_to_{data["end"]}.xlsx"'
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Profit & Loss"

    # تنسيقات
    title_font = Font(bold=True, size=16, color="0f172a")
    bold_font = Font(bold=True)
    red_bold = Font(bold=True, color="dc2626")
    green_bold = Font(bold=True, color="16a34a")
    header_fill = PatternFill("solid", fgColor="f8fafc")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # الترويسة
    ws.merge_cells('A1:D1'); ws['A1'] = 'Trading and Profit & Loss Account'; ws['A1'].font = title_font; ws['A1'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A2:D2'); ws['A2'] = f'For the period {data["period"]}'; ws['A2'].alignment = Alignment(horizontal='center')
    ws.append([])

    # الهيدر
    headers = ['Particulars', 'Amount', 'Particulars', 'Amount']
    ws.append(headers)
    for col in range(1, 5):
        cell = ws.cell(row=4, column=col)
        cell.font = bold_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')

    # بيانات المتاجرة
    ws.append(['Opening Stock', trading['opening_stock'], 'Sales Account', trading['sales']])
    ws.append(['Purchase Account', trading['purchases'], 'Closing Stock', trading['closing_stock']])
    ws.append(['Direct Expenses', trading['direct_expenses'], 'Direct Income', trading['direct_income']])
    ws.append(['Total', trading['opening_stock']+trading['purchases']+trading['direct_expenses'], 'Total', trading['sales']+trading['closing_stock']+trading['direct_income']])
    ws.append(['Gross Profit c/d', trading['gross_profit'], '-', '-'])
    ws.append(['Grand Total', trading['grand_total'], 'Grand Total', trading['grand_total']])
    
    # تلوين الإجماليات باللون الأحمر
    for row in [8, 9, 10]:
        for col in range(1, 5):
            cell = ws.cell(row=row, column=col)
            cell.font = red_bold
            cell.border = thin_border

    # بيانات الربح والخسارة
    ws.append(['-', '-', 'Gross Profit b/f', pl['gross_profit_bf']])
    ws.append(['Indirect Expenses', pl['total_indirect_expenses'], 'Indirect Income', pl['indirect_income']])
    
    for exp in pl['indirect_expenses']:
        ws.append([exp['name'], exp['amount'], '', ''])

    ws.append(['Total', pl['total_indirect_expenses'], 'Total', pl['gross_profit_bf']+pl['indirect_income']])
    ws.append(['Net Profit', pl['net_profit'], '-', '-'])
    ws.append(['Grand Total', pl['grand_total'], 'Grand Total', pl['grand_total']])

    # تلوين الجزء السفلي
    max_r = ws.max_row
    for row in [max_r-2, max_r]: # Total & Grand Total
        for col in range(1, 5):
            ws.cell(row=row, column=col).font = red_bold
    ws.cell(row=max_r-1, column=1).font = green_bold # Net Profit
    ws.cell(row=max_r-1, column=2).font = green_bold

    # إضافة إطارات لكل الخلايا
    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, min_col=1, max_col=4):
        for cell in row:
            cell.border = thin_border

    # ضبط المقاسات
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 15

    wb.save(response)
    return response    

def profit_loss_page(request):
    """ فتح صفحة الربح والخسارة """
    return render(request, 'trading_pl_report.html') # غيرنا الاسم هنا ✅

# ==========================================
# 🛒 تقرير المشتريات (Purchases Report) 🛒
# ==========================================

def purchase_report_page(request):
    """ فتح صفحة تقرير المشتريات """
    return render(request, 'purchase_report.html')

def api_purchase_report_data(request):
    """ API لجلب بيانات المشتريات والموردين مع البحث الذكي """
    try:
        start_date_str = request.GET.get('start', (timezone.now() - timedelta(days=30)).strftime('%Y-%m-%d'))
        end_date_str = request.GET.get('end', timezone.now().strftime('%Y-%m-%d'))
        search_query = request.GET.get('search', '').strip().lower()

        try:
            start_date = datetime.strptime(start_date_str, "%Y-%m-%d").date()
            end_date = datetime.strptime(end_date_str, "%Y-%m-%d").date()
        except ValueError:
            start_date = (timezone.now() - timedelta(days=30)).date()
            end_date = timezone.now().date()

        # جلب المشتريات (المصروفات)
        purchases = Expense.objects.filter(date__gte=start_date, date__lte=end_date)
        
        # البحث برقم الطلب (ID) أو المورد (Description)
        if search_query:
            q_obj = Q(description__icontains=search_query) | Q(notes__icontains=search_query)
            num = ''.join(filter(str.isdigit, search_query))
            if num:
                q_obj |= Q(id=int(num))
            purchases = purchases.filter(q_obj)

        total_purchases = 0.0
        total_paid = 0.0
        completed_count = 0
        transactions_count = purchases.count()
        
        daily_data = {}
        table_data = []

        for p in purchases:
            t_amt = float(p.total_amount or 0.0)
            p_amt = float(p.paid_amount or 0.0)
            balance = t_amt - p_amt
            
            total_purchases += t_amt
            total_paid += p_amt
            
            if balance <= 0.01:
                completed_count += 1
                status = "Paid"
            elif p_amt > 0:
                status = "Partial"
            else:
                status = "Unpaid"

            # تجميع بيانات الشارت
            d_str = p.date.strftime('%Y-%m-%d')
            if d_str not in daily_data:
                daily_data[d_str] = {'amount': 0.0, 'count': 0}
            daily_data[d_str]['amount'] += t_amt
            daily_data[d_str]['count'] += 1

            # تعبئة الجدول
            table_data.append({
                'po_number': f"PO-{p.id:06d}",
                'lpo': p.notes or "-",
                'supplier': p.description or "General Supplier",
                'date': p.date.strftime('%d %b %Y'),
                'total': t_amt,
                'paid': p_amt,
                'balance': balance,
                'status': status
            })

        # ترتيب البيانات تنازلياً للجدول
        table_data.reverse()

        # حساب المؤشرات (KPIs)
        outstanding = total_purchases - total_paid
        outstanding_pct = (outstanding / total_purchases * 100) if total_purchases > 0 else 0.0
        completion_rate = (completed_count / transactions_count * 100) if transactions_count > 0 else 0.0
        days_diff = (end_date - start_date).days + 1
        avg_daily = total_purchases / days_diff if days_diff > 0 else 0.0

        # تجهيز الشارت
        sorted_dates = sorted(daily_data.keys())
        chart_labels = [datetime.strptime(d, '%Y-%m-%d').strftime('%d/%m') for d in sorted_dates]
        chart_amounts = [daily_data[d]['amount'] for d in sorted_dates]
        chart_counts = [daily_data[d]['count'] for d in sorted_dates]

        return JsonResponse({
            "summary": {
                "start": start_date_str,
                "end": end_date_str,
                "total_purchases": round(total_purchases, 2),
                "transactions": transactions_count,
                "outstanding_balance": round(outstanding, 2),
                "outstanding_pct": round(outstanding_pct, 1),
                "completed_pos": completed_count,
                "completion_rate": round(completion_rate, 1),
                "avg_daily": round(avg_daily, 2),
                "days": days_diff,
                "total_paid": round(total_paid, 2)
            },
            "chart": {
                "labels": chart_labels,
                "amounts": chart_amounts,
                "counts": chart_counts
            },
            "table": table_data
        })
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


def export_purchase_excel(request):
    """ تصدير تقرير المشتريات لملف إكسل إبداعي """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    import json

    data_response = api_purchase_report_data(request)
    if data_response.status_code != 200:
        return HttpResponse("Error generating report", status=500)
    
    data = json.loads(data_response.content)
    summary = data['summary']
    table = data['table']

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Purchases_Report_{summary["start"]}_to_{summary["end"]}.xlsx"'
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Purchases Report"

    title_font = Font(bold=True, size=16, color="0f172a")
    header_font = Font(bold=True, color="FFFFFF")
    blue_fill = PatternFill("solid", fgColor="3b82f6")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    ws.append(['AL MARONA LAUNDRY - PURCHASE REPORT (تقرير المشتريات)'])
    ws['A1'].font = title_font
    ws.append([f'Period: {summary["start"]} to {summary["end"]}'])
    ws.append([f'Total Purchases: QAR {summary["total_purchases"]} | Outstanding: QAR {summary["outstanding_balance"]}'])
    ws.append([])

    headers = ['PO Number', 'LPO', 'Supplier', 'Date', 'Total Amount', 'Paid Amount', 'Balance', 'Status']
    ws.append(headers)
    
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=ws.max_row, column=col)
        cell.font = header_font
        cell.fill = blue_fill
        cell.border = thin_border

    for row in table:
        row_data = [row['po_number'], row['lpo'], row['supplier'], row['date'], row['total'], row['paid'], row['balance'], row['status']]
        ws.append(row_data)
        for col in range(1, len(row_data) + 1):
            ws.cell(row=ws.max_row, column=col).border = thin_border

    column_widths = {'A': 15, 'B': 15, 'C': 30, 'D': 15, 'E': 15, 'F': 15, 'G': 15, 'H': 15}
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    wb.save(response)
    return response    

# ==========================================
# 📦 إدارة المخزون والمستهلكات (Inventory) 📦
# ==========================================
from .models import InventoryProduct

def inventory_products_page(request):
    """ فتح صفحة المنتجات والمستهلكات """
    return render(request, 'inventory_products.html')

def api_inventory_products(request):
    """ API لجلب وإضافة المنتجات للمخزون """
    if request.method == 'GET':
        products = InventoryProduct.objects.all().order_by('-created_at')
        data = []
        for p in products:
            data.append({
                'id': p.id,
                'image_url': p.image.url if p.image else None,
                'name_en': p.name_en,
                'name_ar': p.name_ar,
                'sku': p.barcode or "-",
                'category': p.category or "-",
                'brand': p.brand or "-",
                'stock': p.quantity,
                'price': float(p.purchase_price)
            })
        return JsonResponse({'products': data})

    elif request.method == 'POST':
        try:
            # استلام البيانات (بما فيها الصورة)
            p = InventoryProduct(
                barcode=request.POST.get('barcode'),
                name_en=request.POST.get('name_en'),
                name_ar=request.POST.get('name_ar'),
                product_type=request.POST.get('product_type', 'Consumable'),
                quantity=int(request.POST.get('quantity', 0)),
                min_stock_level=int(request.POST.get('min_stock', 0)),
                opening_stock=int(request.POST.get('opening_stock', 0)),
                allow_zero_stock=request.POST.get('allow_zero') == 'true',
                purchase_price=float(request.POST.get('purchase_price', 0.0)),
                unit=request.POST.get('unit'),
                category=request.POST.get('category'),
                brand=request.POST.get('brand'),
                description=request.POST.get('description'),
            )
            
            opening_date = request.POST.get('opening_date')
            if opening_date:
                p.opening_date = datetime.strptime(opening_date, '%Y-%m-%d').date()
                
            if 'image' in request.FILES:
                p.image = request.FILES['image']
                
            p.save()
            return JsonResponse({'message': 'Product saved successfully!'}, status=201)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)


def export_inventory_excel(request):
    """ تصدير قائمة المخزون لإكسل إبداعي """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    products = InventoryProduct.objects.all().order_by('-created_at')

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Inventory_Products_Report.xlsx"'
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventory List"

    title_font = Font(bold=True, size=16, color="ffffff")
    header_font = Font(bold=True, color="0f172a")
    blue_fill = PatternFill("solid", fgColor="0284c7")
    light_fill = PatternFill("solid", fgColor="f1f5f9")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # الترويسة الإبداعية
    ws.merge_cells('A1:G2')
    top_cell = ws['A1']
    top_cell.value = "AL MARONA LAUNDRY - INVENTORY & CONSUMABLES REPORT"
    top_cell.font = title_font
    top_cell.fill = blue_fill
    top_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.append([])

    headers = ['SKU / Barcode', 'Product Name (EN)', 'Product Name (AR)', 'Category', 'Brand', 'Current Stock', 'Unit Cost (QAR)']
    ws.append(headers)
    
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=4, column=col)
        cell.font = header_font
        cell.fill = light_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')

    total_value = 0.0
    for p in products:
        total_value += (p.quantity * float(p.purchase_price))
        row_data = [
            p.barcode or "-", p.name_en, p.name_ar or "-", p.category or "-", 
            p.brand or "-", p.quantity, float(p.purchase_price)
        ]
        ws.append(row_data)
        for col in range(1, len(row_data) + 1):
            cell = ws.cell(row=ws.max_row, column=col)
            cell.border = thin_border
            if col in [6, 7]:
                cell.alignment = Alignment(horizontal='center')

    ws.append([])
    ws.append(['', '', '', '', 'Total Inventory Value:', f"QAR {round(total_value, 2)}"])
    ws.cell(row=ws.max_row, column=5).font = Font(bold=True)
    ws.cell(row=ws.max_row, column=6).font = Font(bold=True, color="dc2626")

    column_widths = {'A': 20, 'B': 30, 'C': 30, 'D': 20, 'E': 20, 'F': 15, 'G': 15}
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    wb.save(response)
    return response    

    # ==========================================
# 📊 تقرير المخزون (Stock Report) 📊
# ==========================================
from django.db.models import Q, F

def stock_report_page(request):
    """ فتح صفحة تقرير المخزون """
    return render(request, 'stock_report.html')

def api_stock_report_data(request):
    """ API لجلب الإحصائيات وبيانات المخزون مع الفلترة """
    try:
        search = request.GET.get('search', '').strip().lower()
        filter_type = request.GET.get('filter', 'all')

        all_products = InventoryProduct.objects.all()

        # 1. حساب المؤشرات العامة (Global KPIs)
        total_products = all_products.count()
        stock_value = sum((p.quantity * float(p.purchase_price)) for p in all_products if p.quantity > 0)
        low_stock_count = all_products.filter(quantity__lte=F('min_stock_level'), quantity__gt=0).count()
        out_of_stock_count = all_products.filter(quantity__lte=0).count()

        # 2. تطبيق الفلاتر والبحث
        products = all_products
        if search:
            products = products.filter(Q(name_en__icontains=search) | Q(name_ar__icontains=search) | Q(barcode__icontains=search))

        if filter_type == 'hardware':
            products = products.filter(product_type__icontains='Hardware')
        elif filter_type == 'consumables':
            products = products.filter(product_type__icontains='Consumable')
        elif filter_type == 'low_stock':
            products = products.filter(quantity__lte=F('min_stock_level'), quantity__gt=0)
        elif filter_type == 'out_of_stock':
            products = products.filter(quantity__lte=0)

        # 3. تجهيز البيانات للواجهة
        data = []
        for p in products:
            status = "In Stock"
            status_color = "text-green-600 bg-green-50"
            if p.quantity <= 0:
                status = "Out of Stock"
                status_color = "text-red-600 bg-red-50"
            elif p.quantity <= p.min_stock_level:
                status = "Low Stock"
                status_color = "text-yellow-600 bg-yellow-50"

            data.append({
                'id': p.id,
                'name_en': p.name_en,
                'name_ar': p.name_ar or "",
                'sku': p.barcode or "-",
                'category': p.category or "-",
                'quantity': p.quantity,
                'price': float(p.purchase_price),
                'value': p.quantity * float(p.purchase_price),
                'status': status,
                'status_color': status_color,
                'image_url': p.image.url if p.image else None
            })

        return JsonResponse({
            'kpis': {
                'total_products': total_products,
                'stock_value': round(stock_value, 2),
                'low_stock': low_stock_count,
                'out_of_stock': out_of_stock_count
            },
            'products': data
        })
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


def export_stock_excel(request):
    """ تصدير تقرير المخزون والجرد لإكسل """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    import json

    data_response = api_stock_report_data(request)
    if data_response.status_code != 200:
        return HttpResponse("Error generating report", status=500)
    
    data = json.loads(data_response.content)
    kpis = data['kpis']
    products = data['products']

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Stock_Report.xlsx"'
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Stock Report"

    title_font = Font(bold=True, size=16, color="0f172a")
    header_font = Font(bold=True, color="FFFFFF")
    blue_fill = PatternFill("solid", fgColor="0284c7")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    ws.append(['AL MARONA LAUNDRY - INVENTORY STOCK REPORT'])
    ws['A1'].font = title_font
    ws.append([f"Total Products: {kpis['total_products']} | Stock Value: QAR {kpis['stock_value']} | Low Stock: {kpis['low_stock']} | Out of Stock: {kpis['out_of_stock']}"])
    ws.append([])

    headers = ['SKU / Barcode', 'Product Name', 'Category', 'Current Stock', 'Unit Cost (QAR)', 'Total Value (QAR)', 'Status']
    ws.append(headers)
    
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=ws.max_row, column=col)
        cell.font = header_font
        cell.fill = blue_fill
        cell.border = thin_border

    for p in products:
        row_data = [p['sku'], p['name_en'], p['category'], p['quantity'], p['price'], p['value'], p['status']]
        ws.append(row_data)
        for col in range(1, len(row_data) + 1):
            cell = ws.cell(row=ws.max_row, column=col)
            cell.border = thin_border
            if p['status'] == 'Out of Stock' and col == 7:
                cell.font = Font(color="dc2626", bold=True)
            elif p['status'] == 'Low Stock' and col == 7:
                cell.font = Font(color="ca8a04", bold=True)

    column_widths = {'A': 20, 'B': 30, 'C': 20, 'D': 15, 'E': 15, 'F': 15, 'G': 15}
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    wb.save(response)
    return response

    # ==========================================
# 💰 تقرير تقييم المخزون المالي (Stock Valuation) 💰
# ==========================================

def inventory_valuation_report_page(request):
    """ فتح صفحة تقييم المخزون (المالية) """
    return render(request, 'inventory_valuation_report.html')

def api_inventory_valuation_data(request):
    """ API لحساب القيمة المالية الحالية للمخزون """
    try:
        from django.utils import timezone
        
        # نجلب فقط المنتجات اللي كميتها أكبر من صفر
        products = InventoryProduct.objects.filter(quantity__gt=0).order_by('-quantity')
        
        total_value = 0.0
        total_items = 0
        categories = set()

        data = []
        for p in products:
            val = p.quantity * float(p.purchase_price)
            total_value += val
            total_items += p.quantity
            cat_name = p.category or 'General'
            categories.add(cat_name)
            
            data.append({
                'name': p.name_en,
                'name_ar': p.name_ar or "",
                'category': cat_name,
                'qty': p.quantity,
                'cost': float(p.purchase_price),
                'total_value': val
            })

        # ترتيب المنتجات تنازلياً حسب القيمة المالية
        data = sorted(data, key=lambda x: x['total_value'], reverse=True)

        return JsonResponse({
            'summary': {
                'total_value': round(total_value, 2),
                'total_items': total_items,
                'product_count': products.count(),
                'categories_count': len(categories),
                'date': timezone.now().strftime('%d %B %Y')
            },
            'products': data
        })
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


def export_inventory_valuation_excel(request):
    """ تصدير تقييم المخزون بملف إكسل للمحاسبين """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    import json
    from django.utils import timezone

    data_response = api_inventory_valuation_data(request)
    if data_response.status_code != 200:
        return HttpResponse("Error generating report", status=500)
    
    data = json.loads(data_response.content)
    summary = data['summary']
    products = data['products']

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Stock_Valuation_{timezone.now().strftime("%Y-%m-%d")}.xlsx"'
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Stock Valuation"

    title_font = Font(bold=True, size=16, color="0f172a")
    header_font = Font(bold=True, color="FFFFFF")
    blue_fill = PatternFill("solid", fgColor="0f172a")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    ws.append(['AL MARONA LAUNDRY - STOCK VALUATION REPORT (تقرير تقييم المخزون)'])
    ws['A1'].font = title_font
    ws.append([f'As of Date: {summary["date"]}'])
    ws.append([f'Total Stock Value: QAR {summary["total_value"]} | Total Unique Products: {summary["product_count"]}'])
    ws.append([])

    headers = ['Product Name', 'Category', 'Quantity in Stock', 'Unit Cost (QAR)', 'Total Value (QAR)']
    ws.append(headers)
    
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=ws.max_row, column=col)
        cell.font = header_font
        cell.fill = blue_fill
        cell.border = thin_border

    for p in products:
        row_data = [p['name'], p['category'], p['qty'], p['cost'], p['total_value']]
        ws.append(row_data)
        for col in range(1, len(row_data) + 1):
            ws.cell(row=ws.max_row, column=col).border = thin_border

    column_widths = {'A': 35, 'B': 20, 'C': 15, 'D': 15, 'E': 20}
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    wb.save(response)
    return response

    # ==========================================
# ⚖️ ميزان المراجعة (Trial Balance) ⚖️
# ==========================================

def trial_balance_page(request):
    """ فتح صفحة ميزان المراجعة """
    return render(request, 'trial_balance.html')

def api_trial_balance_data(request):
    """ API لحساب ميزان المراجعة وتوزين الحسابات (العملاء والموردين) """
    try:
        from django.utils import timezone
        date_str = request.GET.get('date', timezone.now().strftime('%Y-%m-%d'))
        
        try:
            as_of_date = datetime.strptime(date_str, "%Y-%m-%d").date()
        except:
            as_of_date = timezone.now().date()

        # 1. جلب البيانات
        orders = Order.objects.filter(created_at__date__lte=as_of_date, is_void=False)
        payments = Payment.objects.filter(created_at__date__lte=as_of_date)
        expenses = Expense.objects.filter(date__lte=as_of_date)
        inventory = InventoryProduct.objects.filter(quantity__gt=0)

        # 2. حسابات المبيعات والعملاء (Sales & Debtors)
        total_sales = sum(float(o.total_amount or 0) for o in orders)
        cust_cash = sum(float(p.amount or 0) for p in payments if p.method.lower() == 'cash')
        cust_bank = sum(float(p.amount or 0) for p in payments if p.method.lower() != 'cash')
        sundry_debtors = total_sales - (cust_cash + cust_bank) # ذمم مدينة (ديون العملاء)

        # 3. حسابات المشتريات والموردين (Expenses & Accounts Payable)
        total_expenses = sum(float(e.total_amount or 0) for e in expenses)
        paid_expenses = sum(float(e.paid_amount or 0) for e in expenses)
        accounts_payable = total_expenses - paid_expenses # ذمم دائنة (ديون للموردين)

        # 4. حساب المخزون (Inventory)
        inventory_value = sum((p.quantity * float(p.purchase_price)) for p in inventory)

        # 5. بناء الميزان التفصيلي
        detailed = [
            {"account": "STOCK - Opening Stock", "group": "Inventory", "debit": inventory_value, "credit": 0.0},
            {"account": "ASS-0001 - Cash Account", "group": "Cash Account", "debit": cust_cash, "credit": paid_expenses}, # الكاش زاد من الزباين ونقص من الموردين
            {"account": "ASS-0002 - Bank Account (QNB)", "group": "Bank Account", "debit": cust_bank, "credit": 0.0},
            {"account": "ASS-0003 - Sundry Debtors (العملاء)", "group": "Sundry Debtors", "debit": sundry_debtors if sundry_debtors > 0 else 0.0, "credit": abs(sundry_debtors) if sundry_debtors < 0 else 0.0},
            {"account": "INC-0001 - Sales Account", "group": "Sales Account", "debit": 0.0, "credit": total_sales},
            {"account": "EXP-0000 - General Expenses", "group": "Indirect Expense", "debit": total_expenses, "credit": 0.0},
            {"account": "LIA-0001 - Accounts Payable (الموردين)", "group": "Current Liabilities", "debit": 0.0, "credit": accounts_payable},
        ]

        # 6. توزين الميزان (حساب الأرباح أو الخسائر)
        tot_deb = sum(d['debit'] for d in detailed)
        tot_cred = sum(d['credit'] for d in detailed)
        diff = tot_deb - tot_cred
        
        if diff > 0:
            detailed.append({"account": "CAP-0001 - Retained Earnings (Net Profit)", "group": "Capital Account", "debit": 0.0, "credit": diff})
            tot_cred += diff
        elif diff < 0:
            detailed.append({"account": "CAP-0001 - Retained Earnings (Net Loss)", "group": "Capital Account", "debit": abs(diff), "credit": 0.0})
            tot_deb += abs(diff)

        # 7. تجهيز الملخص (Summary)
        summary_dict = {}
        for d in detailed:
            grp = d['group']
            if grp not in summary_dict:
                summary_dict[grp] = {'debit': 0.0, 'credit': 0.0}
            summary_dict[grp]['debit'] += d['debit']
            summary_dict[grp]['credit'] += d['credit']

        summary = [{"group": k, "debit": v['debit'], "credit": v['credit']} for k, v in summary_dict.items()]
        
        detailed.sort(key=lambda x: x['account'])
        summary.sort(key=lambda x: x['group'])

        return JsonResponse({
            "date": date_str,
            "detailed": detailed,
            "summary": summary,
            "total_debit": round(tot_deb, 2),
            "total_credit": round(tot_cred, 2)
        })
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


def export_trial_balance_excel(request):
    """ تصدير ميزان المراجعة (تفصيلي ومجمع) لملف إكسل إبداعي بشيتين """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    import json
    
    data_response = api_trial_balance_data(request)
    if data_response.status_code != 200:
        return HttpResponse("Error generating report", status=500)
    
    data = json.loads(data_response.content)
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Trial_Balance_{data["date"]}.xlsx"'
    wb = openpyxl.Workbook()
    
    title_font = Font(bold=True, size=16, color="0f172a")
    header_font = Font(bold=True, color="FFFFFF")
    blue_fill = PatternFill("solid", fgColor="0284c7")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    def format_sheet(ws, title, is_summary=False):
        ws.append(['AL MARONA LAUNDRY - TRIAL BALANCE (ميزان المراجعة)'])
        ws['A1'].font = title_font
        ws.append([f'As of Date: {data["date"]} | Type: {title}'])
        ws.append([])

        headers = ['Account Group', 'Debit (QAR)', 'Credit (QAR)'] if is_summary else ['Account Name', 'Debit (QAR)', 'Credit (QAR)']
        ws.append(headers)
        
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=4, column=col)
            cell.font = header_font
            cell.fill = blue_fill
            cell.border = thin_border

        items = data['summary'] if is_summary else data['detailed']
        
        for item in items:
            name = item.get('group') if is_summary else item.get('account')
            row_data = [name, item['debit'] or "-", item['credit'] or "-"]
            ws.append(row_data)
            for col in range(1, 4):
                ws.cell(row=ws.max_row, column=col).border = thin_border

        # Totals
        ws.append(['Total', data['total_debit'], data['total_credit']])
        for col in range(1, 4):
            c = ws.cell(row=ws.max_row, column=col)
            c.font = Font(bold=True)
            c.border = thin_border

        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15

    # الشيت الأول (التفصيلي)
    ws1 = wb.active
    ws1.title = "Detailed"
    format_sheet(ws1, "Detailed Trial Balance", is_summary=False)

    # الشيت الثاني (المجمع)
    ws2 = wb.create_sheet(title="Summary")
    format_sheet(ws2, "Summary Trial Balance", is_summary=True)

    wb.save(response)
    return response

# ==========================================
# 📖 تصدير كتالوج أصناف الغسيل مع الصور والأسعار المستعجلة (Service Catalog) 📖
# ==========================================
def export_service_catalog_excel(request):
    import openpyxl
    from openpyxl.drawing.image import Image as OpenpyxlImage
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    import os
    from django.http import HttpResponse
    from django.utils import timezone
    from services.models import LaundryItem

    # 1. إنشاء ملف الإكسل
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Services Catalog"

    # 2. التنسيقات الفخمة 
    title_font = Font(bold=True, size=18, color="1e293b")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    blue_fill = PatternFill("solid", fgColor="0284c7")
    light_gray_fill = PatternFill("solid", fgColor="f8fafc")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # 🌟 حل مشكلة الترويسة العلوية (دمج محكم مع لون وحدود) 🌟
    ws.merge_cells('A1:H2')
    top_cell = ws['A1']
    top_cell.value = "AL MARONA LAUNDRY - SERVICES & PRICES CATALOG"
    top_cell.font = title_font
    top_cell.alignment = center_align
    top_cell.fill = light_gray_fill
    
    # تطبيق الإطار على الترويسة عشان ما تطلع الخلية الأولى مفصولة
    for row in ws['A1:H2']:
        for cell in row:
            cell.border = thin_border

    # 🌟 حل مشكلة عناوين الأعمدة (إجبارها على سطرين) 🌟
    headers = [
        'Image\n(الصورة)', 
        'Item Name\n(الصنف)', 
        'Ironing\n(كوي)', 
        'Washing & Ironing\n(غسيل وكوي)', 
        'Dry Clean\n(دراي كلين)', 
        'Express Ironing\n(كوي مستعجل)', 
        'Express Wash & Iron\n(غسيل وكوي مستعجل)', 
        'Express Dry Clean\n(دراي كلين مستعجل)'
    ]
    
    # تكبير ارتفاع صف العناوين (عشان النص الطويل ياخذ راحته)
    ws.row_dimensions[3].height = 45 
    
    for col_num, header_title in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num)
        cell.value = header_title
        cell.font = header_font
        cell.fill = blue_fill
        cell.alignment = center_align
        cell.border = thin_border

    # عرض الأعمدة
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 22
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 26
    ws.column_dimensions['H'].width = 22

    # 3. جلب البيانات من الداتابيز
    items = LaundryItem.objects.all().order_by('category', 'name_en')

    row_num = 4 # البيانات تبدأ من الصف الرابع
    for item in items:
        ws.row_dimensions[row_num].height = 80 

        # جلب الأسعار بطريقة آمنة
        iron_price = f"{float(item.ironing_price)} QAR" if getattr(item, 'ironing_price', 0) else "-"
        wash_price = f"{float(item.washing_price)} QAR" if getattr(item, 'washing_price', 0) else "-"
        dry_price = f"{float(item.dry_cleaning_price)} QAR" if getattr(item, 'dry_cleaning_price', 0) else "-"
        
        iron_exp = f"{float(item.ironing_express_price)} QAR" if getattr(item, 'ironing_express_price', 0) else "-"
        wash_exp = f"{float(item.washing_express_price)} QAR" if getattr(item, 'washing_express_price', 0) else "-"
        dry_exp = f"{float(item.dry_cleaning_express_price)} QAR" if getattr(item, 'dry_cleaning_express_price', 0) else "-"

        ws.cell(row=row_num, column=2, value=f"{item.name_en or ''}\n{item.name or ''}").alignment = center_align
        ws.cell(row=row_num, column=3, value=iron_price).alignment = center_align
        ws.cell(row=row_num, column=4, value=wash_price).alignment = center_align
        ws.cell(row=row_num, column=5, value=dry_price).alignment = center_align
        ws.cell(row=row_num, column=6, value=iron_exp).alignment = center_align
        ws.cell(row=row_num, column=7, value=wash_exp).alignment = center_align
        ws.cell(row=row_num, column=8, value=dry_exp).alignment = center_align

        # معالجة وتركيب الصورة
        if item.image and hasattr(item.image, 'path') and os.path.exists(item.image.path):
            try:
                img = OpenpyxlImage(item.image.path)
                img.width = 90  
                img.height = 90 
                ws.add_image(img, f'A{row_num}') 
            except Exception as e:
                ws.cell(row=row_num, column=1, value="Img Error").alignment = center_align
        else:
            ws.cell(row=row_num, column=1, value="No Image").alignment = center_align

        for col in range(1, 9):
            ws.cell(row=row_num, column=col).border = thin_border

        row_num += 1

    # 4. إرسال الكتالوج
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Laundry_Catalog_{timezone.now().strftime("%Y-%m-%d")}.xlsx"'
    wb.save(response)
    return response